"""Importa cadastros a partir de uma planilha .xlsx (varias abas).

Regras:
  - mapeia ABA POR ABA (cada aba tem seu proprio cabecalho);
  - o MAC e detectado pelo PADRAO (regex), nao pela posicao -> resolve abas
    onde MAC e Nome estao trocados;
  - a UNIDADE vem do NOME DA ABA (de-para abaixo);
  - colunas que nao se aplicam vao para 'observacoes' (notes);
  - um MAC pode aparecer em varias abas -> os dados sao mesclados.
"""
from __future__ import annotations

import os
import re
import unicodedata

from openpyxl import load_workbook

from .client import UnifiClient, UnifiError

MAC_RE = re.compile(r"(?i)\b([0-9a-f]{2}[:\-]){5}[0-9a-f]{2}\b")

# Cabecalho normalizado -> campo do cadastro
HEADER_SYN = {
    "mac": "mac",
    "nome": "nome", "name": "nome",
    "setor": "setor",
    "gestor": "lider", "lider": "lider",
    "chamado": "chamado",
    "funcao": "funcao", "cargo": "funcao",
    "unidade": "_unidade_col",  # tratado a parte
}

# Palavra-chave no nome da aba -> numero da unidade.
# O mapa REAL (nomes das suas unidades) fica em "sites_map.json" na raiz do
# projeto (arquivo LOCAL, fora do versionamento). Se nao existir, usa o exemplo
# generico abaixo. Formato do JSON: {"sheet_unit": {"nome da aba": "101", ...},
#                                     "valid_units": ["101","102", ...]}
_DEFAULT_SHEET_UNIT = {"exemplo unidade a": "101", "exemplo unidade b": "102"}
_DEFAULT_VALID_UNITS = {"101", "102", "103", "104", "105"}


def _load_sites_map():
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.environ.get("SITES_MAP") or os.path.join(base, "sites_map.json")
    try:
        import json
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        su = {str(k).lower(): str(v) for k, v in (data.get("sheet_unit") or {}).items()}
        vu = {str(x) for x in (data.get("valid_units") or [])}
        return (su or _DEFAULT_SHEET_UNIT, vu or _DEFAULT_VALID_UNITS)
    except Exception:
        return (_DEFAULT_SHEET_UNIT, _DEFAULT_VALID_UNITS)


SHEET_UNIT, VALID_UNITS = _load_sites_map()


def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.lower()


def unit_for_sheet(sheet_name: str) -> str | None:
    n = _norm(sheet_name)
    for kw, unit in SHEET_UNIT.items():
        if kw in n:
            return unit
    return None


def _find_mac(cells) -> str | None:
    for c in cells:
        if c is None:
            continue
        m = MAC_RE.search(str(c))
        if m:
            try:
                return UnifiClient.normalize_mac(m.group(0))
            except UnifiError:
                continue
    return None


def parse_workbook(path: str) -> tuple[dict, list[dict]]:
    """Retorna (records_por_mac, stats_por_aba)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    records: dict[str, dict] = {}
    stats: list[dict] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        unit = unit_for_sheet(sheet)
        rows = ws.iter_rows(values_only=True)
        header = None
        for r in rows:
            if r and any(v is not None and str(v).strip() for v in r):
                header = list(r)
                break
        if not header:
            stats.append({"sheet": sheet, "unit": unit, "rows": 0,
                          "imported": 0, "skipped": 0})
            continue

        # mapeia colunas
        col_field = {}   # idx -> campo
        col_label = {}   # idx -> rotulo p/ notes
        for i, h in enumerate(header):
            key = _norm(h)
            if not key:
                continue
            if key in HEADER_SYN:
                col_field[i] = HEADER_SYN[key]
            else:
                col_label[i] = str(h).strip()

        n_rows = n_imp = n_skip = 0
        for r in rows:
            if not r or not any(v is not None and str(v).strip() for v in r):
                continue
            n_rows += 1
            cells = list(r)
            mac = _find_mac(cells)
            if not mac:
                n_skip += 1
                continue
            n_imp += 1

            rec = records.setdefault(mac, {
                "mac": mac, "nome": "", "setor": "", "funcao": "",
                "lider": "", "chamado": "", "unidades": set(), "notes": [],
                "sheets": set(),
            })
            rec["sheets"].add(sheet)
            if unit:
                rec["unidades"].add(unit)

            leftover_name = ""  # caso MAC/Nome trocados
            for i, val in enumerate(cells):
                if val is None or not str(val).strip():
                    continue
                sval = str(val).strip()
                # e o proprio MAC desta linha? entao ignora (ja capturado)
                mm = MAC_RE.search(sval)
                if mm:
                    try:
                        if UnifiClient.normalize_mac(mm.group(0)) == mac:
                            continue
                    except UnifiError:
                        pass
                field = col_field.get(i)
                if field == "mac":
                    # header "MAC" mas o valor nao e um MAC (colunas trocadas)
                    leftover_name = leftover_name or sval
                    continue
                if field == "_unidade_col":
                    for tok in re.split(r"[,;/ ]+", sval):
                        if tok in VALID_UNITS:
                            rec["unidades"].add(tok)
                        elif tok:
                            rec["notes"].append(f"[{sheet}] Unidade: {tok}")
                    continue
                if field in ("nome", "setor", "funcao", "lider", "chamado"):
                    if not rec[field]:
                        rec[field] = sval
                    elif rec[field] != sval:
                        rec["notes"].append(f"[{sheet}] {field}: {sval}")
                    continue
                # coluna nao mapeada -> observacoes
                label = col_label.get(i, "col")
                rec["notes"].append(f"[{sheet}] {label}: {sval}")

            if not rec["nome"] and leftover_name:
                rec["nome"] = leftover_name

        stats.append({"sheet": sheet, "unit": unit, "rows": n_rows,
                      "imported": n_imp, "skipped": n_skip})

    return records, stats


def finalize(rec: dict) -> dict:
    """Converte um record interno em campos do client_info."""
    return {
        "nome": rec["nome"],
        "setor": rec["setor"],
        "funcao": rec["funcao"],
        "lider": rec["lider"],
        "chamado": rec["chamado"],
        "unidade": ", ".join(sorted(rec["unidades"])),
        "notes": " | ".join(rec["notes"])[:2000],
    }
